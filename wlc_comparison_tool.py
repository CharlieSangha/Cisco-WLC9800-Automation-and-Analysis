#!/usr/bin/env python3
import os #Required for computer folders and files.
import re #Finds patterns in text.
import pandas as pd #organise data in table.
import glob #find files that match certain patterns.
from pathlib import Path #Another helper for working with folders and files.
import argparse #Gives instructions when running the script. 
import logging #Writes down messages on what the script is doing.
from openpyxl import Workbook #Creates the Excel File. 
from openpyxl.styles import PatternFill, Font
import openpyxl

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Create command line arguments
parser = argparse.ArgumentParser(description='Analyze and compare Cisco WLC data')
parser.add_argument('--base-dir', default='./device_info', 
                    help='Base directory containing device subdirectories (default: ./device_info)')
parser.add_argument('--output', default='wlc_comparison_report.xlsx',
                    help='Output filename (default: wlc_comparison_report.xlsx)')
args = parser.parse_args()

# Create a function to find all device directories
def find_device_directories(base_path):
    """
    Find all device directories under the specified base path.
    
    Args:
        base_path (str): Path to the base directory containing device subdirectories
    
    Returns:
        list: List of device directory paths
    """
    logger.info(f"Looking for device directories in {base_path}")
    
    # Make sure the base directory exists
    if not os.path.exists(base_path):
        logger.error(f"Base directory {base_path} does not exist")
        return []
    
    # Find all subdirectories in the base directory
    device_dirs = [d for d in os.listdir(base_path) 
                  if os.path.isdir(os.path.join(base_path, d))]
    
    logger.info(f"Found {len(device_dirs)} device directories")
    return [os.path.join(base_path, d) for d in device_dirs]

# Create parsers for each file type
def parse_hostname(file_path):
    """
    Parse hostname from hostname.txt file.
    
    Args:
        file_path (str): Path to the hostname.txt file
    
    Returns:
        str: Hostname of the device
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read().strip()
        
        # Extract hostname from "hostname XXXXX" format
        match = re.search(r'hostname\s+(\S+)', content)
        if match:
            return match.group(1)
        return content
    except Exception as e:
        logger.error(f"Error parsing hostname from {file_path}: {e}")
        return "Unknown"

def parse_version(file_path):
    """
    Parse version information from version.txt file.
    
    Args:
        file_path (str): Path to the version.txt file
    
    Returns:
        dict: Dictionary containing version information
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        version_info = {}
        
        # Extract Cisco IOS XE Software version
        ios_xe_match = re.search(r'Cisco IOS XE Software, Version\s+([^\s,]+)', content)
        if ios_xe_match:
            version_info['ios_xe_version'] = ios_xe_match.group(1)
        
        # Extract System image file
        system_image_match = re.search(r'System image file is "([^"]+)"', content)
        if system_image_match:
            version_info['system_image'] = system_image_match.group(1)
        
        # Extract model
        model_match = re.search(r'cisco\s+(\S+)(?:\s+\(|\s+\()([^)]+)\)', content)
        if model_match:
            version_info['model'] = model_match.group(1)
            version_info['processor'] = model_match.group(2)
        
        return version_info
    except Exception as e:
        logger.error(f"Error parsing version from {file_path}: {e}")
        return {}

def parse_wlan_summary(file_path):
    """
    Parse WLAN information from wlan.txt file.
    
    Args:
        file_path (str): Path to the wlan.txt file
    
    Returns:
        dict: Dictionary containing WLAN information
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        # If content shows "No WLAN configuration found", return empty values
        if "No WLAN configuration found" in content:
            return {"wlan_count": 0, "wlans": []}
        
        wlan_info = {}
        wlans = []
        
        # Log the first part of the content for debugging
        logger.debug(f"WLAN content sample: {content[:500]}")
        
        # More flexible pattern to match different Cisco WLC WLAN summary formats
        # First, try the common format
        wlan_entries = re.findall(r'^\s*(\d+)\s+(\S+\s*\S*)\s+(\S+)\s+', content, re.MULTILINE)
        
        if not wlan_entries:
            # If no matches, try alternative format (adjust based on actual output)
            logger.info("First WLAN pattern didn't match, trying alternative format")
            wlan_entries = re.findall(r'^\s*(\d+)\s+([^\s]+[^\n]*?)\s+(\S+)\s+', content, re.MULTILINE)
        
        # Process each entry with flexible auth type extraction
        for entry in wlan_entries:
            wlan_id = entry[0]
            name = entry[1].strip()
            status = entry[2]
            
            # Extract auth type more flexibly (may appear in different positions)
            auth_match = re.search(r'\[(.*?)\]', content[content.find(name):content.find(name) + 200])
            auth = auth_match.group(1).strip() if auth_match else "N/A"
            
            wlans.append({
                'id': wlan_id,
                'name': name,
                'status': status,
                'auth': auth
            })
        
        # If still no entries, log the content for troubleshooting
        if not wlans:
            logger.warning(f"Failed to parse any WLANs. Content sample: {content[:1000]}")
        
        wlan_info['wlan_count'] = len(wlans)
        wlan_info['wlans'] = wlans
        
        return wlan_info
    except Exception as e:
        logger.error(f"Error parsing WLAN summary from {file_path}: {e}")
        return {"wlan_count": 0, "wlans": []}

def parse_ap_summary(file_path):
    """
    Parse AP summary information from ap_summary.txt file.
    
    Args:
        file_path (str): Path to the ap_summary.txt file
    
    Returns:
        dict: Dictionary containing AP summary information
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        # If content shows "No AP summary found", return empty dict
        if "No AP summary found" in content:
            return {"ap_count": 0, "ap_up": 0, "ap_down": 0, "ap_models": {}}
        
        ap_info = {}
        ap_models = {}
        
        # Count total APs and status
        ap_entries = re.findall(r'^\S+\s+\S+\s+\S+\s+(\S+)', content, re.MULTILINE)
        
        total_aps = len(ap_entries)
        up_aps = ap_entries.count('Registered')
        down_aps = total_aps - up_aps
        
        # Count AP models
        model_matches = re.findall(r'(\S+)\s+\S+\s+\S+\s+\S+', content, re.MULTILINE)
        for model in model_matches:
            if model not in ap_models:
                ap_models[model] = 1
            else:
                ap_models[model] += 1
        
        ap_info['ap_count'] = total_aps
        ap_info['ap_up'] = up_aps
        ap_info['ap_down'] = down_aps
        ap_info['ap_models'] = ap_models
        
        return ap_info
    except Exception as e:
        logger.error(f"Error parsing AP summary from {file_path}: {e}")
        return {"ap_count": 0, "ap_up": 0, "ap_down": 0, "ap_models": {}}

def parse_tag_site(file_path):
    """
    Parse site tag information from tag_site.txt file.
    
    Args:
        file_path (str): Path to the tag_site.txt file
    
    Returns:
        list: List of site tags
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        # If content shows no tag site found, return empty list
        if "No wireless tag site summary found" in content:
            return []
        
        # Extract number of site tags
        count_match = re.search(r'Number of Site Tags:\s*(\d+)', content)
        site_tag_count = int(count_match.group(1)) if count_match else 0
        
        # Skip header rows and extract actual tag entries
        content_lines = content.strip().split('\n')
        
        # Find the line with "Tag Name" and "Description" (the header row)
        header_line_index = -1
        for i, line in enumerate(content_lines):
            if "Site Tag Name" in line and "Description" in line:
                header_line_index = i
                break
        
        # If we found a header line, look for actual tag entries after it and the separator line
        tags = []
        if header_line_index >= 0 and len(content_lines) > header_line_index + 2:
            # Skip header and separator line (usually dashes)
            for line in content_lines[header_line_index + 2:]:
                if not line.strip():
                    continue
                
                # Split line by whitespace, but keep the first part (tag name)
                # and combine the rest as description
                parts = line.strip().split(None, 1)
                if len(parts) >= 1:
                    name = parts[0]
                    desc = parts[1] if len(parts) > 1 else ""
                    tags.append({'name': name, 'description': desc})
        
        return tags
    except Exception as e:
        logger.error(f"Error parsing tag site from {file_path}: {e}")
        return []

def parse_vlan_brief(file_path):
    """
    Parse VLAN information from vlan_brief.txt file.
    
    Args:
        file_path (str): Path to the vlan_brief.txt file
    
    Returns:
        list: List of VLANs
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        # If content shows no vlan found, return empty list
        if "No vlan configuration found" in content:
            return []
        
        # Extract VLAN entries
        vlan_entries = re.findall(r'^\s*(\d+)\s+(\S+)\s+(\S+)', content, re.MULTILINE)
        
        # Filter out header row and dashes
        vlans = [{'id': int(vlan_id), 'name': name, 'status': status} 
                for vlan_id, name, status in vlan_entries 
                if vlan_id.isdigit() and name != '----']
        
        return vlans
    except Exception as e:
        logger.error(f"Error parsing VLAN brief from {file_path}: {e}")
        return []

def parse_ntp(file_path):
    """
    Parse NTP server information from ntp.txt file.
    
    Args:
        file_path (str): Path to the ntp.txt file
    
    Returns:
        list: List of NTP servers
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        # If content shows no NTP found, return empty list
        if "No NTP configuration found" in content:
            return []
        
        # Extract NTP server entries
        ntp_servers = re.findall(r'ntp server\s+(\S+)', content, re.MULTILINE)
        
        # Clean up any unexpected values like "ip"
        cleaned_servers = []
        for server in ntp_servers:
            # Skip entries that just say "ip" or are not valid hostnames/IPs
            if server.lower() == "ip" or not server.strip():
                continue
            cleaned_servers.append(server)
        
        return cleaned_servers
    except Exception as e:
        logger.error(f"Error parsing NTP from {file_path}: {e}")
        return []

def parse_aaa_servers(file_path):
    """
    Parse AAA server information from aaa_servers.txt file.
    
    Args:
        file_path (str): Path to the aaa_servers.txt file
    
    Returns:
        list: List of AAA servers
    """
    try:
        with open(file_path, 'r') as f:
            content = f.read()
        
        # If content shows no AAA servers found, return empty list
        if "No AAA servers found" in content:
            return []
        
        # Extract AAA server entries - more comprehensive pattern for Cisco outputs
        aaa_servers = []
        
        # Look for RADIUS server entries in the format shown in the example
        radius_entries = re.findall(r'RADIUS:\s+id\s+(\d+).*?host\s+(\S+),\s+auth-port\s+(\d+).*?hostname\s+(\S+).*?State:\s+current\s+(\S+)', 
                                    content, re.DOTALL)
        
        for entry in radius_entries:
            server_id, ip, auth_port, hostname, status = entry
            aaa_servers.append({
                'id': server_id,
                'ip': ip,
                'auth_port': auth_port,
                'hostname': hostname,
                'status': status
            })
        
        # If the above pattern doesn't match, try a more general pattern
        if not aaa_servers:
            server_sections = re.split(r'RADIUS:', content)
            for section in server_sections:
                if not section.strip():
                    continue
                
                # Extract IP address
                ip_match = re.search(r'host\s+(\d+\.\d+\.\d+\.\d+)', section)
                if not ip_match:
                    continue
                
                ip = ip_match.group(1)
                
                # Extract other details
                auth_port_match = re.search(r'auth-port\s+(\d+)', section)
                auth_port = auth_port_match.group(1) if auth_port_match else "N/A"
                
                hostname_match = re.search(r'hostname\s+(\S+)', section)
                hostname = hostname_match.group(1) if hostname_match else ip
                
                status_match = re.search(r'State:\s+current\s+(\S+)', section)
                status = status_match.group(1) if status_match else "Unknown"
                
                aaa_servers.append({
                    'ip': ip,
                    'auth_port': auth_port,
                    'hostname': hostname,
                    'status': status
                })
        
        return aaa_servers
    except Exception as e:
        logger.error(f"Error parsing AAA servers from {file_path}: {e}")
        return []

def process_device_directory(device_dir):
    """
    Process a single device directory to extract all information.
    
    Args:
        device_dir (str): Path to the device directory
    
    Returns:
        dict: Dictionary containing all parsed information
    """
    logger.info(f"Processing device directory: {device_dir}")
    
    device_data = {}
    device_name = os.path.basename(device_dir)
    device_data['directory_name'] = device_name
    
    # Process each file
    hostname_file = os.path.join(device_dir, 'hostname.txt')
    if os.path.exists(hostname_file):
        device_data['hostname'] = parse_hostname(hostname_file)
    else:
        device_data['hostname'] = device_name
    
    version_file = os.path.join(device_dir, 'version.txt')
    if os.path.exists(version_file):
        device_data.update(parse_version(version_file))
    
    wlan_file = os.path.join(device_dir, 'wlan.txt')
    if os.path.exists(wlan_file):
        wlan_info = parse_wlan_summary(wlan_file)
        device_data['wlan_count'] = wlan_info['wlan_count']
        device_data['wlans'] = wlan_info['wlans']
    
    ap_file = os.path.join(device_dir, 'ap_summary.txt')
    if os.path.exists(ap_file):
        ap_info = parse_ap_summary(ap_file)
        device_data.update(ap_info)
    
    tag_site_file = os.path.join(device_dir, 'tag_site.txt')
    if os.path.exists(tag_site_file):
        device_data['site_tags'] = parse_tag_site(tag_site_file)
    
    vlan_file = os.path.join(device_dir, 'vlan_brief.txt')
    if os.path.exists(vlan_file):
        device_data['vlans'] = parse_vlan_brief(vlan_file)
    
    ntp_file = os.path.join(device_dir, 'ntp.txt')
    if os.path.exists(ntp_file):
        device_data['ntp_servers'] = parse_ntp(ntp_file)
    
    aaa_file = os.path.join(device_dir, 'aaa_servers.txt')
    if os.path.exists(aaa_file):
        device_data['aaa_servers'] = parse_aaa_servers(aaa_file)
    
    logger.info(f"Completed processing for device: {device_data['hostname']}")
    return device_data

def highlight_differences(df):
    """
    Highlight differences in a DataFrame.
    Returns a styled DataFrame with differences highlighted.
    
    Args:
        df (pandas.DataFrame): The DataFrame to highlight differences in
    
    Returns:
        pandas.io.formats.style.Styler: Styled DataFrame with differences highlighted
    """
    def highlight_diff(s):
        if s.nunique() > 1:
            return ['background-color: #FFCC99' for _ in s]
        else:
            return ['background-color: white' for _ in s]
    
    return df.style.apply(highlight_diff, axis=1)

def create_excel_report(devices_data, output_file):
    """
    Create an Excel report with multiple sheets for different aspects of the comparison.
    
    Args:
        devices_data (list): List of dictionaries containing device data
        output_file (str): Output Excel file path
    """
    logger.info(f"Creating Excel report: {output_file}")
    
    # Create a new workbook
    wb = Workbook()
    
    # Create overview sheet
    overview_sheet = wb.active
    overview_sheet.title = "Overview"
    
    # Set up headers for overview
    headers = ["Device", "Hostname", "IOS XE Version", "Model", "AP Count", "AP Up", "AP Down", "WLAN Count"]
    for col_num, header in enumerate(headers, 1):
        cell = overview_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add device overview data
    for row_num, device in enumerate(devices_data, 2):
        overview_sheet.cell(row=row_num, column=1).value = device.get('directory_name', 'Unknown')
        overview_sheet.cell(row=row_num, column=2).value = device.get('hostname', 'Unknown')
        overview_sheet.cell(row=row_num, column=3).value = device.get('ios_xe_version', 'Unknown')
        overview_sheet.cell(row=row_num, column=4).value = device.get('model', 'Unknown')
        overview_sheet.cell(row=row_num, column=5).value = device.get('ap_count', 0)
        overview_sheet.cell(row=row_num, column=6).value = device.get('ap_up', 0)
        overview_sheet.cell(row=row_num, column=7).value = device.get('ap_down', 0)
        overview_sheet.cell(row=row_num, column=8).value = device.get('wlan_count', 0)
    
    # Highlight differences in the overview sheet
    diff_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    
    # Check each column for differences
    for col_num in range(3, 9):  # Skip device name and hostname columns
        values = set()
        for row_num in range(2, len(devices_data) + 2):
            cell_value = overview_sheet.cell(row=row_num, column=col_num).value
            values.add(str(cell_value))  # Convert to string for comparison
        
        # If there are differences, highlight the cells
        if len(values) > 1:
            for row_num in range(2, len(devices_data) + 2):
                overview_sheet.cell(row=row_num, column=col_num).fill = diff_fill
    
    # Create WLANs sheet
    wlan_sheet = wb.create_sheet(title="WLANs")
    
    # Set up headers for WLANs
    wlan_headers = ["Device", "WLAN ID", "WLAN Name", "Status", "Auth Type"]
    for col_num, header in enumerate(wlan_headers, 1):
        cell = wlan_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add WLAN data
    row_num = 2
    for device in devices_data:
        hostname = device.get('hostname', 'Unknown')
        wlans = device.get('wlans', [])
        
        # If we have no WLANs but there should be some based on the count
        if not wlans and device.get('wlan_count', 0) > 0:
            logger.warning(f"Device {hostname} has {device.get('wlan_count')} WLANs but none were parsed properly")
        
        for wlan in wlans:
            wlan_sheet.cell(row=row_num, column=1).value = hostname
            wlan_sheet.cell(row=row_num, column=2).value = wlan.get('id', 'Unknown')
            wlan_sheet.cell(row=row_num, column=3).value = wlan.get('name', 'Unknown')
            wlan_sheet.cell(row=row_num, column=4).value = wlan.get('status', 'Unknown')
            wlan_sheet.cell(row=row_num, column=5).value = wlan.get('auth', 'Unknown')
            row_num += 1
        
    # Check if we added any WLANs
    if row_num == 2:
        # No WLANs were added, add a note
        for col_num in range(1, 6):
            wlan_sheet.cell(row=2, column=col_num).value = "No WLAN data found"
    
    # Create VLANs sheet
    vlan_sheet = wb.create_sheet(title="VLANs")
    
    # Set up headers for VLANs
    vlan_headers = ["Device", "VLAN ID", "VLAN Name", "Status"]
    for col_num, header in enumerate(vlan_headers, 1):
        cell = vlan_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add VLAN data
    row_num = 2
    for device in devices_data:
        hostname = device.get('hostname', 'Unknown')
        vlans = device.get('vlans', [])
        
        for vlan in vlans:
            vlan_sheet.cell(row=row_num, column=1).value = hostname
            vlan_sheet.cell(row=row_num, column=2).value = vlan.get('id', 'Unknown')
            vlan_sheet.cell(row=row_num, column=3).value = vlan.get('name', 'Unknown')
            vlan_sheet.cell(row=row_num, column=4).value = vlan.get('status', 'Unknown')
            row_num += 1
    
    # Create NTP sheet
    ntp_sheet = wb.create_sheet(title="NTP Servers")
    
    # Set up headers for NTP
    ntp_headers = ["Device", "NTP Server"]
    for col_num, header in enumerate(ntp_headers, 1):
        cell = ntp_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add NTP data
    row_num = 2
    for device in devices_data:
        hostname = device.get('hostname', 'Unknown')
        ntp_servers = device.get('ntp_servers', [])
        
        # If we have NTP servers, add them
        if ntp_servers:
            for server in ntp_servers:
                ntp_sheet.cell(row=row_num, column=1).value = hostname
                ntp_sheet.cell(row=row_num, column=2).value = server
                row_num += 1
        else:
            # If no NTP servers, add a row with "None configured"
            ntp_sheet.cell(row=row_num, column=1).value = hostname
            ntp_sheet.cell(row=row_num, column=2).value = "None configured"
            row_num += 1
    
    # Check if we added any NTP servers
    if row_num == 2:
        # No NTP servers were added, add a note
        for col_num in range(1, 3):
            ntp_sheet.cell(row=2, column=col_num).value = "No NTP data found"
    
    # Create AAA Servers sheet
    aaa_sheet = wb.create_sheet(title="AAA Servers")
    
    # Set up headers for AAA
    aaa_headers = ["Device", "Server IP", "Auth Port", "Hostname", "Status"]
    for col_num, header in enumerate(aaa_headers, 1):
        cell = aaa_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add AAA data
    row_num = 2
    for device in devices_data:
        hostname = device.get('hostname', 'Unknown')
        aaa_servers = device.get('aaa_servers', [])
        
        if aaa_servers:
            for server in aaa_servers:
                aaa_sheet.cell(row=row_num, column=1).value = hostname
                aaa_sheet.cell(row=row_num, column=2).value = server.get('ip', 'Unknown')
                aaa_sheet.cell(row=row_num, column=3).value = server.get('auth_port', 'Unknown')
                aaa_sheet.cell(row=row_num, column=4).value = server.get('hostname', 'Unknown')
                aaa_sheet.cell(row=row_num, column=5).value = server.get('status', 'Unknown')
                row_num += 1
        else:
            # If no AAA servers, add a row with "None configured"
            aaa_sheet.cell(row=row_num, column=1).value = hostname
            aaa_sheet.cell(row=row_num, column=2).value = "None configured"
            row_num += 1
    
    # Check if we added any AAA servers
    if row_num == 2:
        # No AAA servers were added, add a note
        aaa_sheet.cell(row=2, column=1).value = "No AAA data found"
    
    # Create Site Tags sheet
    tag_sheet = wb.create_sheet(title="Site Tags")
    
    # Set up headers for Site Tags
    tag_headers = ["Device", "Tag Name", "Description"]
    for col_num, header in enumerate(tag_headers, 1):
        cell = tag_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add Site Tag data
    row_num = 2
    for device in devices_data:
        hostname = device.get('hostname', 'Unknown')
        site_tags = device.get('site_tags', [])
        
        if site_tags:
            for tag in site_tags:
                tag_sheet.cell(row=row_num, column=1).value = hostname
                tag_sheet.cell(row=row_num, column=2).value = tag.get('name', 'Unknown')
                tag_sheet.cell(row=row_num, column=3).value = tag.get('description', 'Unknown')
                row_num += 1
        else:
            # If no Site Tags, add a row with "None configured"
            tag_sheet.cell(row=row_num, column=1).value = hostname
            tag_sheet.cell(row=row_num, column=2).value = "None configured"
            row_num += 1
    
    # Check if we added any site tags
    if row_num == 2:
        # No site tags were added, add a note
        tag_sheet.cell(row=2, column=1).value = "No Site Tag data found"
    
    # Save the workbook
    wb.save(output_file)
    logger.info(f"Excel report saved to {output_file}")

def main():
    """
    Main function to run the script.
    """
    logger.info("Starting Cisco WLC Comparison Tool")
    
    # Find all device directories
    device_dirs = find_device_directories(args.base_dir)
    
    if not device_dirs:
        logger.error(f"No device directories found in {args.base_dir}")
        return
    
    # Process each device directory
    all_device_data = []
    for device_dir in device_dirs:
        device_data = process_device_directory(device_dir)
        all_device_data.append(device_data)
    
    # Create Excel report
    create_excel_report(all_device_data, args.output)
    
    logger.info("Analysis complete!")
    logger.info(f"Report saved to: {args.output}")

if __name__ == "__main__":
    main()
